import openpyxl

wb = openpyxl.load_workbook("Book1.xlsx")
# ワークシート名
print(wb.sheetnames)

# 内容が入っているセルの範囲
sheet = wb.worksheets[0]
print(sheet.dimensions)

# 行数と列数の読み取り
print(sheet.max_row, sheet.max_column)

# セルの値を読み取り
print(sheet.cell(1,1).value)
print(sheet['A1'].value)
print(sheet['A1:C2'])

# すべての値にアクセス
for row in range(1, sheet.max_row + 1):
    for col in "ABC":
        print(sheet[f"{col}{row}"].value, end="\t")
    print()